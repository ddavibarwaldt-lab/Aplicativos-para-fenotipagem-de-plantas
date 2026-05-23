import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import openpyxl
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm

OPCOES_CARACTERISTICAS = {
    "Código da amostra": [],
    "Hábito de crescimento": ["Ausente", "Ereto", "Semi-erecto", "Deitado", "Prostrado", "Semi-arrosetado", "Arrosetado"],
    "Forma da folha": ["Ausente", "Inteira", "Lobulada", "Dissectada"],
    "Cor do talo": ["Ausente", "Verde", "Verde com poucas manchas", "Verde com muitas manchas",
                    "Pigmentado com abundante verde", "Pigmentado com pouco verde",
                    "Avermelhado", "Roxo"],
    "Forma das asas do talo": ["Ausente", "Retas", "Onduladas", "Dentadas"],
    "Forma da corola": ["Ausente", "Estrelada", "Semi-estrelada", "Pentagonal", "Rotada", "Muito rotada"],
    "Cor predominante da flor": ["Ausente", "Branco", "Vermelho-rosado", "Vermelho-roxo", "Celeste", "Azul-roxo", "Lilás", "Roxo", "Violeta"],
    "Intensidade da cor da flor": ["Ausente", "Pálido / Claro", "Intermediário", "Intenso / Escuro"],
    "Cor secundária da flor": ["Ausente", "Branco", "Vermelho-rosado", "Vermelho-roxo", "Celeste", "Azul-roxo", "Lilás", "Roxo", "Violeta"],
    "Distribuição da cor secundária da flor": [
        "Ausente", "Ponta (frente)", "Ponta (verso)", "Ponta (ambos)",
        "Em estrela", "Faixas na frente", "Faixas no verso", "Faixas em ambos", "Manchas salpicadas"
    ],
    "V.- Grau de Floração": [
        "0 sem botões", "Presença", "Ausência", "1 Abortamento de botões",
        "3 Floração escassa", "5 Floração moderada", "7 Floração abundante"
    ],
    "VIII.- Pigmentação nas Anteras (Fig. 8)": [
        "1", "0 Sem antocianinas", "1 Faixas laterais pigmentadas",
        "2 Mancha pigmentada no ápice", "3 Faixas e ápice pigmentados",
        "Presença", "Ausência", "4 Anteras vermelho-marrom"
    ],
    "IX.- Pigmentação no Pistilo (Fig. 9)": [
        "Presença", "Ausência", "0 Sem antocianinas", "1 Estigma pigmentado",
        "2 Ovário pigmentado", "3 Pigmento na parede interna do ovário",
        "4 Pigmentado Estigma+Ovário", "5 Pigmentado Estigma+Parede interna",
        "6 Pigmentado Ovário+Parede interna", "7 Pigmentado Estigma+Ovário+Parede interna"
    ],
    "X.- Cor do Cálice": [
        "1 Verde", "Predominantemente pigmentado", "Predominantemente verde", "2 Verde com poucas manchas",
        "3 Verde com muitas manchas", "4 Pigmentado com abundante verde",
        "5 Pigmentado com pouco verde", "6 Avermelhado", "7 Roxo"
    ],
    "XI.- Cor do Pedicelo": [
        "1 Verde", "2 Apenas articulação pigmentada", "3 Ligeiramente pigmentado ao longo sem articulação",
        "4 Pigmentação ao longo e na articulação", "5 Pigmentado sobre a articulação",
        "6 Pigmentado abaixo da articulação", "7 Predominantemente pigmentado e articulação verde",
        "Bastante pigmentado", "Bastante verde", "8 Completamente pigmentado"
    ],
    "Forma geral do tubérculo": ["Ausente", "Comprimido", "Redondo", "Ovalado", "Obovado", "Elíptico", "Oblongo", "Oblongo-alargado", "Alargado"],
    "Variante de forma do tubérculo": ["Ausente", "Aplanado", "Clavado", "Reniforme", "Fusiforme", "Falcado", "Enroscado", "Digitado", "Concertinado", "Tuberosado"],
    "Profundidade dos olhos": ["Ausente", "Sobressalente", "Superficial", "Médio", "Profundo", "Muito profundo"],
    "Cor da pele do tubérculo": ["Ausente", "Branco-creme", "Amarelo", "Laranja", "Marrom", "Rosado", "Vermelho", "Vermelho-roxo", "Roxo", "Negro"],
    "Intensidade da cor da pele": ["Ausente", "Pálido / Claro", "Intermediário", "Intenso / Escuro"],
    "Cor secundária da pele": ["Ausente", "Branco-creme", "Amarelo", "Rosado", "Vermelho", "Vermelho-roxo", "Roxo", "Negro"],
    "Distribuição da cor secundária da pele": ["Ausente", "Nos olhos", "Nas sobrancelhas", "Ao redor dos olhos", "Manchas dispersas", "Como óculos", "Manchas salpicadas", "Poucas manchas"],
    "Cor da polpa do tubérculo": ["Ausente", "Branco", "Creme", "Amarelo claro", "Amarelo", "Amarelo intenso", "Vermelho", "Roxo", "Violeta"],
    "Cor secundária da polpa": ["Ausente", "Branco", "Creme", "Amarelo claro", "Amarelo", "Vermelho", "Roxo", "Violeta"],
    "Distribuição da cor secundária da polpa": [
        "Ausente", "Poucas manchas", "Áreas", "Anel vascular estreito", "Anel vascular largo",
        "Anel vascular e medula", "Todo menos medula", "Outro (salpicado)"
    ],
    "Comprimento da folha 1 (cm)": [],
    "Número de foliolos folha 1": [],
    "Número interfolíolos folha 1": [],
    "Área da folha 1 (cm²)": [],
    "Comprimento da folha 2 (cm)": [],
    "Número de foliolos folha 2": [],
    "Número interfolíolos folha 2": [],
    "Área da folha 2 (cm²)": [],
    "Comprimento da folha 3 (cm)": [],
    "Número de foliolos folha 3": [],
    "Número interfolíolos folha 3": [],
    "Área da folha 3 (cm²)": []
}

# --- A classe Aplicacao e funções permanecem iguais ---
class Aplicacao:
    def __init__(self, master):
        self.master = master
        master.title("Caracterização de Batata")
        master.geometry("1000x700")

        self.entries = {}
        self.caminho_excel = None

        main_frame = tk.Frame(master)
        main_frame.pack(fill=tk.BOTH, expand=1)

        canvas_form = tk.Canvas(main_frame)
        scrollbar_form = ttk.Scrollbar(main_frame, orient=tk.VERTICAL, command=canvas_form.yview)
        self.form_frame = tk.Frame(canvas_form)

        self.form_frame.bind("<Configure>", lambda e: canvas_form.configure(scrollregion=canvas_form.bbox("all")))

        canvas_form.create_window((0, 0), window=self.form_frame, anchor="nw")
        canvas_form.configure(yscrollcommand=scrollbar_form.set)

        canvas_form.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
        scrollbar_form.pack(side=tk.RIGHT, fill=tk.Y)

        for i, (chave, opcoes) in enumerate(OPCOES_CARACTERISTICAS.items()):
            label = tk.Label(self.form_frame, text=chave)
            label.grid(row=i, column=0, sticky="w", padx=5, pady=3)
            if opcoes:
                cb = ttk.Combobox(self.form_frame, values=opcoes, width=40)
                cb.grid(row=i, column=1, sticky="w", padx=5, pady=3)
                cb.current(0)
                self.entries[chave] = cb
            else:
                entry = tk.Entry(self.form_frame, width=42)
                entry.grid(row=i, column=1, sticky="w", padx=5, pady=3)
                self.entries[chave] = entry

        btn_frame = tk.Frame(master)
        btn_frame.pack(fill=tk.X, padx=10, pady=10)

        tk.Button(btn_frame, text="Abrir Imagem", command=self.abrir_imagem).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Salvar dados", command=self.salvar_dados).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Exportar Excel", command=self.exportar_excel).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="Carregar Tabela", command=self.carregar_excel_existente).pack(side=tk.LEFT, padx=5)

        self.frame_imagem = tk.Frame(master, bd=2, relief=tk.SUNKEN)
        self.frame_imagem.pack(fill=tk.BOTH, expand=1, padx=10, pady=10)

        self.canvas_imagem = tk.Canvas(self.frame_imagem, bg="gray")
        self.canvas_imagem.pack(side=tk.LEFT, fill=tk.BOTH, expand=1)

        self.scrollbar_vertical = ttk.Scrollbar(self.frame_imagem, orient=tk.VERTICAL, command=self.canvas_imagem.yview)
        self.scrollbar_vertical.pack(side=tk.RIGHT, fill=tk.Y)

        self.scrollbar_horizontal = ttk.Scrollbar(self.frame_imagem, orient=tk.HORIZONTAL, command=self.canvas_imagem.xview)
        self.scrollbar_horizontal.pack(side=tk.BOTTOM, fill=tk.X)

        self.canvas_imagem.configure(yscrollcommand=self.scrollbar_vertical.set, xscrollcommand=self.scrollbar_horizontal.set)
        self.canvas_imagem.bind('<Configure>', self.redimensionar_canvas_imagem)

        self.imagem_tk = None
        self.img_id = None

    def redimensionar_canvas_imagem(self, event):
        self.canvas_imagem.configure(scrollregion=self.canvas_imagem.bbox("all"))

    def abrir_imagem(self):
        caminho = filedialog.askopenfilename(
            title="Selecionar imagem",
            filetypes=[("Arquivos de imagem", "*.png;*.jpg;*.jpeg;*.bmp;*.gif")]
        )
        if caminho:
            imagem = Image.open(caminho)
            self.imagem_tk = ImageTk.PhotoImage(imagem)
            self.canvas_imagem.delete("all")
            self.img_id = self.canvas_imagem.create_image(0, 0, anchor="nw", image=self.imagem_tk)
            self.canvas_imagem.config(scrollregion=self.canvas_imagem.bbox(tk.ALL))

    def salvar_dados(self):
        dados = {chave: widget.get() for chave, widget in self.entries.items()}
        messagebox.showinfo("Dados", f"Dados salvos:\n{dados}")

    def carregar_excel_existente(self):
        caminho = filedialog.askopenfilename(
            title="Abrir arquivo Excel existente",
            filetypes=[("Arquivos Excel", "*.xlsx")]
        )
        if caminho:
            self.caminho_excel = caminho
            messagebox.showinfo("Tabela carregada", f"Arquivo carregado:\n{os.path.basename(caminho)}")

    def exportar_excel(self):
        if not self.caminho_excel:
            self.caminho_excel = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Arquivos Excel", "*.xlsx")],
                title="Salvar arquivo Excel"
            )
            if not self.caminho_excel:
                return

        try:
            if os.path.exists(self.caminho_excel):
                wb = openpyxl.load_workbook(self.caminho_excel)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                col = 1
                for chave in self.entries.keys():
                    ws.cell(row=1, column=col, value=chave)
                    col += 1

            nova_linha = ws.max_row + 1
            col = 1
            for chave, widget in self.entries.items():
                ws.cell(row=nova_linha, column=col, value=widget.get())
                col += 1

            wb.save(self.caminho_excel)
            messagebox.showinfo("Sucesso", "Dados salvos com sucesso na tabela!")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao salvar Excel: {e}")

def main():
    root = tk.Tk()
    app = Aplicacao(root)
    root.mainloop()

if __name__ == "__main__":
    main()
