import tkinter as tk
from tkinter import filedialog, colorchooser, messagebox, simpledialog
from PIL import Image, ImageTk
import cv2
import numpy as np
import os
import math
from datetime import datetime

# openpyxl é opcional — se não instalado o botão de salvar ficará desativado
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class GreenAreaDetector:
    def __init__(self, root):
        self.root = root
        self.root.title("FENOTIPAGEM")

        # Frame principal com canvas e scrollbars
        self.frame = tk.Frame(root)
        self.frame.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(self.frame, bg="black")
        self.h_scroll = tk.Scrollbar(self.frame, orient=tk.HORIZONTAL, command=self.canvas.xview)
        self.v_scroll = tk.Scrollbar(self.frame, orient=tk.VERTICAL, command=self.canvas.yview)

        self.canvas.configure(xscrollcommand=self.h_scroll.set, yscrollcommand=self.v_scroll.set)
        self.canvas.grid(row=0, column=0, sticky='nsew')
        self.h_scroll.grid(row=1, column=0, sticky='ew')
        self.v_scroll.grid(row=0, column=1, sticky='ns')

        self.frame.grid_rowconfigure(0, weight=1)
        self.frame.grid_columnconfigure(0, weight=1)

        # Caixa de informações rápida
        self.info_box = tk.Text(root, height=6, width=100)
        self.info_box.pack(padx=5, pady=4)

        # Caixa de resultados contínuos (lista + resumo)
        self.results_box = tk.Text(root, height=12, width=100)
        self.results_box.pack(padx=5, pady=4)

        # Botões
        buttons_frame = tk.Frame(root)
        buttons_frame.pack(pady=5)

        open_btn = tk.Button(buttons_frame, text="Abrir Imagem", command=self.select_image)
        open_btn.grid(row=0, column=0, padx=5)

        reset_btn = tk.Button(buttons_frame, text="Resetar", command=self.reset_all)
        reset_btn.grid(row=0, column=1, padx=5)

        measure_btn = tk.Button(buttons_frame, text="Medir Distância", command=self.activate_distance_mode)
        measure_btn.grid(row=0, column=2, padx=5)

        color_btn = tk.Button(buttons_frame, text="Selecionar Cor", command=self.choose_color)
        color_btn.grid(row=0, column=3, padx=5)

        save_btn = tk.Button(buttons_frame, text="Salvar XLSX", command=self.save_results)
        save_btn.grid(row=0, column=4, padx=5)

        clear_results_btn = tk.Button(buttons_frame, text="Limpar Resultados", command=self.clear_results)
        clear_results_btn.grid(row=0, column=5, padx=5)

        # Botão de tolerância
        tolerance_btn = tk.Button(buttons_frame, text="Tolerância (%)", command=self.set_tolerance)
        tolerance_btn.grid(row=0, column=6, padx=5)

        # desativa salvar se openpyxl faltando
        if not HAS_OPENPYXL:
            save_btn.config(state=tk.DISABLED)
            self.info_box.insert(tk.END, "openpyxl não encontrado: instale 'openpyxl' para permitir salvar em XLSX.\n")

        # Variáveis
        self.original_img = None  # numpy RGB
        self.display_img = None
        self.tk_img = None
        self.img_id = None

        self.scale_points = []
        self.rect_start = None
        self.regions = []
        self.region_rect_ids = []

        self.distance_mode = False
        self.distance_points = []

        # Cor padrão (violeta forte)
        self.selected_color = (148, 0, 211)

        # tolerância %
        self.tolerance_percent = 0.0

        # Armazena resultados: lista de dicionários
        self.results_data = []

        # Menu
        menu = tk.Menu(root)
        root.config(menu=menu)
        file_menu = tk.Menu(menu, tearoff=0)
        menu.add_cascade(label="Arquivo", menu=file_menu)
        file_menu.add_command(label="Abrir Imagem", command=self.select_image)

        # binds
        self.canvas.bind("<Button-1>", self.on_click)
        self.canvas.bind("<B1-Motion>", self.on_drag)
        self.canvas.bind("<ButtonRelease-1>", self.on_release)

        # instrução inicial
        self.info_box.insert(tk.END, "1. Clique em dois pontos para definir a escala (px -> cm).\n")

    # ---------------- UI helpers ----------------
    def choose_color(self):
        color_code = colorchooser.askcolor(title="Escolha a cor para destacar pixels")
        if color_code[0] is not None:
            r, g, b = map(int, color_code[0])
            self.selected_color = (r, g, b)
            self.info_box.insert(tk.END, f"Cor escolhida: RGB {self.selected_color}\n")

    def set_tolerance(self):
        try:
            val = simpledialog.askfloat("Tolerância (%)", "Insira a tolerância percentual (ex: 10 para 10%):", minvalue=0.0)
            if val is None:
                return
            self.tolerance_percent = float(val)
            self.info_box.insert(tk.END, f"Tolerância definida: {self.tolerance_percent:.2f}%\n")
        except Exception as e:
            self.info_box.insert(tk.END, f"Valor inválido de tolerância: {e}\n")

    def reset_all(self):
        self.scale_points = []
        self.rect_start = None
        self.regions = []
        # apagar retângulos visuais no canvas
        for rid in getattr(self, 'region_rect_ids', []):
            try:
                self.canvas.delete(rid)
            except Exception:
                pass
        self.region_rect_ids = []
        self.pixels_per_cm = None
        self.distance_mode = False
        self.distance_points = []
        self.info_box.delete("1.0", tk.END)
        self.info_box.insert(tk.END, "1. Clique em dois pontos para definir a escala\n")
        self.display_img = self.original_img.copy() if self.original_img is not None else None
        self.update_image_on_canvas()

    def clear_results(self):
        self.results_data = []
        self.update_results_box()
        self.info_box.insert(tk.END, "Resultados limpos.\n")

    def select_image(self):
        folder_path = filedialog.askdirectory(title="Escolha a pasta com imagens")
        if not folder_path:
            return

        images = [f for f in os.listdir(folder_path) if f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))]
        if not images:
            self.info_box.insert(tk.END, "Nenhuma imagem encontrada na pasta.\n")
            return

        popup = tk.Toplevel(self.root)
        popup.title("Escolher Imagem")

        thumbs_frame = tk.Frame(popup)
        thumbs_frame.pack(fill="both", expand=True)

        canvas = tk.Canvas(thumbs_frame)
        scroll_y = tk.Scrollbar(thumbs_frame, orient="vertical", command=canvas.yview)
        scroll_frame = tk.Frame(canvas)

        scroll_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=scroll_frame, anchor="nw")
        canvas.configure(yscrollcommand=scroll_y.set)

        canvas.pack(side="left", fill="both", expand=True)
        scroll_y.pack(side="right", fill="y")

        self.thumbs = []
        for i, img_file in enumerate(images):
            img_path = os.path.join(folder_path, img_file)
            try:
                pil_img = Image.open(img_path).resize((200, 150))
                tk_thumb = ImageTk.PhotoImage(pil_img)
                self.thumbs.append(tk_thumb)

                frame = tk.Frame(scroll_frame, bd=2, relief="ridge", padx=5, pady=5)
                frame.grid(row=i//3, column=i%3, padx=10, pady=10)

                btn = tk.Button(frame, image=tk_thumb, command=lambda p=img_path: self.load_image_from_popup(popup, p))
                btn.pack()

                label = tk.Label(frame, text=img_file, wraplength=180)
                label.pack()
            except Exception as e:
                print(f"Erro ao carregar {img_file}: {e}")

    def load_image_from_popup(self, popup, path):
        popup.destroy()
        self.load_image(path)

    def load_image(self, file_path):
        bgr = cv2.imread(file_path)
        if bgr is None:
            messagebox.showerror("Erro", f"Não foi possível abrir {file_path}")
            return
        self.original_img = cv2.cvtColor(bgr, cv2.COLOR_BGR2RGB)
        self.display_img = self.original_img.copy()
        self.scale_points = []
        self.regions = []
        self.region_rect_ids = []
        self.rect_start = None
        self.pixels_per_cm = None
        self.distance_mode = False
        self.distance_points = []
        self.info_box.delete("1.0", tk.END)
        self.info_box.insert(tk.END, "1. Clique em dois pontos para definir a escala\n")
        self.update_image_on_canvas()

    def update_image_on_canvas(self):
        if self.display_img is not None:
            pil = Image.fromarray(self.display_img)
            self.tk_img = ImageTk.PhotoImage(pil)
            self.canvas.delete("all")
            self.img_id = self.canvas.create_image(0, 0, anchor="nw", image=self.tk_img)
            self.canvas.config(scrollregion=self.canvas.bbox(tk.ALL))

    # ---------------- eventos do canvas ----------------
    def on_click(self, event):
        x = int(self.canvas.canvasx(event.x))
        y = int(self.canvas.canvasy(event.y))

        if self.distance_mode:
            self.distance_points.append((x, y))
            self.canvas.create_oval(x-3, y-3, x+3, y+3, fill='yellow', outline='')
            if len(self.distance_points) == 2:
                self.calculate_distance()
            return

        # escala (dois pontos)
        if len(self.scale_points) < 2:
            self.scale_points.append((x, y))
            self.canvas.create_oval(x-3, y-3, x+3, y+3, fill='red', outline='')
            if len(self.scale_points) == 2:
                self.ask_scale()
        elif len(self.regions) < 2:
            self.rect_start = (x, y)

    def on_drag(self, event):
        if self.rect_start and len(self.regions) < 2:
            x = int(self.canvas.canvasx(event.x))
            y = int(self.canvas.canvasy(event.y))

            # remove último retângulo temporário
            if self.region_rect_ids:
                try:
                    self.canvas.delete(self.region_rect_ids[-1])
                except Exception:
                    pass

            x0, y0 = self.rect_start
            rect_id = self.canvas.create_rectangle(x0, y0, x, y, outline='blue')
            if len(self.region_rect_ids) == len(self.regions):
                self.region_rect_ids.append(rect_id)
            else:
                self.region_rect_ids[-1] = rect_id

    def on_release(self, event):
        if self.rect_start and len(self.regions) < 2:
            x0, y0 = self.rect_start
            x1 = int(self.canvas.canvasx(event.x))
            y1 = int(self.canvas.canvasy(event.y))

            x0, y0, x1, y1 = map(int, (min(x0, x1), min(y0, y1), max(x0, x1), max(y0, y1)))
            h, w = self.original_img.shape[:2]
            x0, x1 = max(0, x0), min(w, x1)
            y0, y1 = max(0, y0), min(h, y1)

            # evita regiões vazias
            if x1 - x0 <= 0 or y1 - y0 <= 0:
                self.info_box.insert(tk.END, "Região inválida (muito pequena). Tente novamente.\n")
                self.rect_start = None
                return

            self.regions.append((x0, y0, x1, y1))
            self.rect_start = None

            if len(self.regions) == 1:
                # Após selecionar o primeiro retângulo (tom mais escuro)
                self.info_box.insert(tk.END, "2. Selecione a região do tom mais claro (arraste para selecionar).\n")
            elif len(self.regions) == 2:
                # duas regiões selecionadas -> calcula
                self.info_box.insert(tk.END, "3. Calculando área com base nos tons selecionados...\n")
                self.root.after(80, self.calculate_green_area)

    # ---------------- caixa de escala ----------------
    def ask_scale(self):
        popup = tk.Toplevel(self.root)
        popup.title("Definir Escala")
        tk.Label(popup, text="Distância real entre os pontos (em cm):").pack(padx=10, pady=6)
        scale_entry = tk.Entry(popup)
        scale_entry.pack(padx=10, pady=6)

        def confirm():
            try:
                real_distance = float(scale_entry.get())
                (px1, py1), (px2, py2) = self.scale_points
                pixel_distance = math.dist((px1, py1), (px2, py2))
                if real_distance <= 0:
                    raise ValueError("Distância real precisa ser > 0")
                self.pixels_per_cm = pixel_distance / real_distance  # px por cm
                popup.destroy()
                self.info_box.insert(tk.END, f"Escala definida: {self.pixels_per_cm:.3f} px/cm\n")
                self.info_box.insert(tk.END, "Agora selecione a região do tom mais escuro (arraste para selecionar).\n")
            except Exception as e:
                self.info_box.insert(tk.END, f"Valor inválido: {e}\n")

        tk.Button(popup, text="OK", command=confirm).pack(pady=6)

    # ---------------- medição de distância ----------------
    def activate_distance_mode(self):
        if not hasattr(self, 'pixels_per_cm') or not self.pixels_per_cm:
            self.info_box.insert(tk.END, "Defina a escala antes de medir distância.\n")
            return
        self.distance_mode = True
        self.distance_points = []
        self.info_box.insert(tk.END, "Modo de medição ativado: clique em dois pontos na imagem.\n")

    def calculate_distance(self):
        if len(self.distance_points) != 2:
            return
        (x1, y1), (x2, y2) = self.distance_points
        pixel_dist = math.dist((x1, y1), (x2, y2))
        real_dist_cm = pixel_dist / self.pixels_per_cm
        real_dist_cm_rounded = round(real_dist_cm, 2)

        self.info_box.insert(tk.END, f"Distância real medida: {real_dist_cm_rounded:.2f} cm\n")

        entry = {
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Tipo": "Distância",
            "Comprimento_cm": real_dist_cm_rounded,
            "Área_cm2": None,
            "Pixels": None
        }
        self.results_data.append(entry)
        self.update_results_box()

        self.distance_mode = False
        self.distance_points = []

    # ---------------- cálculo de área verde (ajustado para permitir múltiplas medições) ----------------
    def calculate_green_area(self):
        try:
            if not hasattr(self, 'pixels_per_cm') or not self.pixels_per_cm:
                self.info_box.insert(tk.END, "Erro: Escala não definida corretamente.\n")
                return

            if len(self.regions) < 2:
                self.info_box.insert(tk.END, "Erro: Duas regiões precisam ser selecionadas.\n")
                return

            hsv = cv2.cvtColor(self.original_img.copy(), cv2.COLOR_RGB2HSV)

            def extract_region_hsv(rect):
                x0, y0, x1, y1 = rect
                roi = hsv[y0:y1, x0:x1]
                if roi.size == 0:
                    raise ValueError("Região vazia")
                h_min = int(np.min(roi[:, :, 0]))
                s_min = int(np.min(roi[:, :, 1]))
                v_min = int(np.min(roi[:, :, 2]))
                h_max = int(np.max(roi[:, :, 0]))
                s_max = int(np.max(roi[:, :, 1]))
                v_max = int(np.max(roi[:, :, 2]))
                return (np.array([h_min, s_min, v_min], dtype=np.int32),
                        np.array([h_max, s_max, v_max], dtype=np.int32))

            # regiões: [0] = tom mais escuro, [1] = tom mais claro
            hsv_min_raw, _ = extract_region_hsv(self.regions[0])  # tom mais escuro (valores menores)
            _, hsv_max_raw = extract_region_hsv(self.regions[1])  # tom mais claro (valores maiores)

            # calculo do centro e meio-intervalo
            hsv_min_raw = hsv_min_raw.astype(np.int32)
            hsv_max_raw = hsv_max_raw.astype(np.int32)
            center = ((hsv_min_raw + hsv_max_raw) / 2.0).astype(np.float32)
            half_range = ((hsv_max_raw - hsv_min_raw) / 2.0).astype(np.float32)

            # aplica tolerância: aumenta o half_range por (1 + tol%)
            tol_factor = 1.0 + (self.tolerance_percent / 100.0)
            new_half = half_range * tol_factor

            hsv_lower = (center - new_half).astype(np.int32)
            hsv_upper = (center + new_half).astype(np.int32)

            # clip para intervalos válidos do HSV em OpenCV: H:0-179, S:0-255, V:0-255
            hsv_lower[0] = max(0, min(179, hsv_lower[0]))
            hsv_lower[1] = max(0, min(255, hsv_lower[1]))
            hsv_lower[2] = max(0, min(255, hsv_lower[2]))
            hsv_upper[0] = max(0, min(179, hsv_upper[0]))
            hsv_upper[1] = max(0, min(255, hsv_upper[1]))
            hsv_upper[2] = max(0, min(255, hsv_upper[2]))

            hsv_lower_u8 = hsv_lower.astype(np.uint8)
            hsv_upper_u8 = hsv_upper.astype(np.uint8)

            # --- NOVO: extrair RGB min/max das regiões selecionadas para exibir e salvar ---
            rgb_img = self.original_img.copy()

            def extract_region_rgb(rect):
                x0, y0, x1, y1 = rect
                roi = rgb_img[y0:y1, x0:x1]
                if roi.size == 0:
                    raise ValueError("Região vazia RGB")
                r_min = int(np.min(roi[:, :, 0]))
                g_min = int(np.min(roi[:, :, 1]))
                b_min = int(np.min(roi[:, :, 2]))
                r_max = int(np.max(roi[:, :, 0]))
                g_max = int(np.max(roi[:, :, 1]))
                b_max = int(np.max(roi[:, :, 2]))
                return (r_min, g_min, b_min), (r_max, g_max, b_max)

            rgb_min_dark, rgb_max_dark = extract_region_rgb(self.regions[0])
            rgb_min_light, rgb_max_light = extract_region_rgb(self.regions[1])

            # Máscara com os limites ajustados
            mask = cv2.inRange(hsv, hsv_lower_u8, hsv_upper_u8)
            green_pixels = int(np.count_nonzero(mask))

            result_img = self.display_img.copy() if self.display_img is not None else self.original_img.copy()
            # pinta com a cor escolhida (RGB)
            result_img[mask > 0] = self.selected_color

            # atualiza display e mantém original_img intacta (permitindo novas medições)
            self.display_img = result_img
            self.update_image_on_canvas()

            pixel_area_cm2 = (1 / self.pixels_per_cm) ** 2
            total_area_cm2 = green_pixels * pixel_area_cm2

            # Converte os limites HSV lower/upper em RGB para exibir
            hsv_lower_pix = np.uint8([[[hsv_lower_u8[0], hsv_lower_u8[1], hsv_lower_u8[2]]]])
            hsv_upper_pix = np.uint8([[[hsv_upper_u8[0], hsv_upper_u8[1], hsv_upper_u8[2]]]])
            rgb_lower_pix = cv2.cvtColor(hsv_lower_pix, cv2.COLOR_HSV2RGB)[0, 0].astype(int)
            rgb_upper_pix = cv2.cvtColor(hsv_upper_pix, cv2.COLOR_HSV2RGB)[0, 0].astype(int)

            # Exibe na info_box — HSV originais e ajustados, e RGB das regiões (min/max)
            orig_lower = np.minimum(hsv_min_raw, hsv_max_raw).astype(np.int32)
            orig_upper = np.maximum(hsv_min_raw, hsv_max_raw).astype(np.int32)
            orig_lower_u8 = np.clip(orig_lower, [0, 0, 0], [179, 255, 255]).astype(np.uint8)
            orig_upper_u8 = np.clip(orig_upper, [0, 0, 0], [179, 255, 255]).astype(np.uint8)
            orig_rgb_lower = cv2.cvtColor(np.uint8([[[orig_lower_u8[0], orig_lower_u8[1], orig_lower_u8[2]]]]), cv2.COLOR_HSV2RGB)[0,0].astype(int)
            orig_rgb_upper = cv2.cvtColor(np.uint8([[[orig_upper_u8[0], orig_upper_u8[1], orig_upper_u8[2]]]]), cv2.COLOR_HSV2RGB)[0,0].astype(int)

            self.info_box.insert(tk.END, f"Pixels detectados: {green_pixels}\n")
            self.info_box.insert(tk.END, f"Área estimada: {total_area_cm2:.2f} cm²\n\n")

            self.info_box.insert(tk.END, "Intervalos HSV (originais calculados a partir das regiões):\n")
            self.info_box.insert(tk.END, f"  HSV min original: H={orig_lower[0]} S={orig_lower[1]} V={orig_lower[2]}\n")
            self.info_box.insert(tk.END, f"  HSV max original: H={orig_upper[0]} S={orig_upper[1]} V={orig_upper[2]}\n")
            self.info_box.insert(tk.END, f"  RGB min correspondente (orig): R={int(orig_rgb_lower[0])} G={int(orig_rgb_lower[1])} B={int(orig_rgb_lower[2])}\n")
            self.info_box.insert(tk.END, f"  RGB max correspondente (orig): R={int(orig_rgb_upper[0])} G={int(orig_rgb_upper[1])} B={int(orig_rgb_upper[2])}\n\n")

            self.info_box.insert(tk.END, f"Tolerância aplicada: {self.tolerance_percent:.2f}%\n")
            self.info_box.insert(tk.END, "Intervalos HSV (após aplicação da tolerância):\n")
            self.info_box.insert(tk.END, f"  HSV lower (ajustado): H={int(hsv_lower_u8[0])} S={int(hsv_lower_u8[1])} V={int(hsv_lower_u8[2])}\n")
            self.info_box.insert(tk.END, f"  HSV upper (ajustado): H={int(hsv_upper_u8[0])} S={int(hsv_upper_u8[1])} V={int(hsv_upper_u8[2])}\n")
            self.info_box.insert(tk.END, f"  RGB lower correspondente: R={int(rgb_lower_pix[0])} G={int(rgb_lower_pix[1])} B={int(rgb_lower_pix[2])}\n")
            self.info_box.insert(tk.END, f"  RGB upper correspondente: R={int(rgb_upper_pix[0])} G={int(rgb_upper_pix[1])} B={int(rgb_upper_pix[2])}\n\n")

            # Exibe RGB min/max das regiões selecionadas
            self.info_box.insert(tk.END, "Valores RGB das regiões selecionadas:\n")
            self.info_box.insert(tk.END, f"  Tom Escuro - RGB min: {rgb_min_dark}  RGB max: {rgb_max_dark}\n")
            self.info_box.insert(tk.END, f"  Tom Claro  - RGB min: {rgb_min_light}  RGB max: {rgb_max_light}\n\n")

            entry = {
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Tipo": "Área (tom selecionado)",
                "Comprimento_cm": None,
                "Área_cm2": round(total_area_cm2, 2),
                "Pixels": green_pixels,
                # dados adicionais
                "HSV_lower": tuple(int(x) for x in hsv_lower_u8),
                "HSV_upper": tuple(int(x) for x in hsv_upper_u8),
                "RGB_lower_from_HSV": tuple(int(x) for x in rgb_lower_pix),
                "RGB_upper_from_HSV": tuple(int(x) for x in rgb_upper_pix),
                "RGB_min_escuro": rgb_min_dark,
                "RGB_max_escuro": rgb_max_dark,
                "RGB_min_claro": rgb_min_light,
                "RGB_max_claro": rgb_max_light,
                "Tolerance_percent": float(self.tolerance_percent)
            }
            self.results_data.append(entry)
            self.update_results_box()

            # limpa seleção para nova medição mantendo a mesma escala
            for rid in getattr(self, 'region_rect_ids', []):
                try:
                    self.canvas.delete(rid)
                except Exception:
                    pass
            self.region_rect_ids = []
            self.regions = []
            self.rect_start = None

            self.info_box.insert(tk.END, "Medição concluída. Para nova medição na mesma imagem, selecione novamente o tom mais escuro.\n")

        except Exception as e:
            self.info_box.insert(tk.END, f"Erro ao calcular área: {e}\n")

    # ---------------- atualização da caixa de resultados (lista + resumo) ----------------
    def update_results_box(self):
        self.results_box.delete("1.0", tk.END)
        header = f"{'#':<4}{'Tipo':<25}{'Comprimento (cm)':<20}{'Área (cm²)':<15}{'Pixels':<10}{'Hora':<20}\n"
        self.results_box.insert(tk.END, header)
        self.results_box.insert(tk.END, "-" * 110 + "\n")
        for i, r in enumerate(self.results_data, start=1):
            comp = f"{r['Comprimento_cm']:.2f}" if r['Comprimento_cm'] is not None else ""
            area = f"{r['Área_cm2']:.2f}" if r['Área_cm2'] is not None else ""
            pixels = f"{r['Pixels']}" if r['Pixels'] is not None else ""
            line = f"{i:<4}{r['Tipo']:<25}{comp:<20}{area:<15}{pixels:<10}{r['timestamp']:<20}\n"
            self.results_box.insert(tk.END, line)

        # resumo estatístico (médias)
        self.results_box.insert(tk.END, "\nResumo:\n")
        # médias
        len_vals = [r['Comprimento_cm'] for r in self.results_data if r['Comprimento_cm'] is not None]
        area_vals = [r['Área_cm2'] for r in self.results_data if r['Área_cm2'] is not None]
        pixel_vals = [r['Pixels'] for r in self.results_data if r['Pixels'] is not None]

        def mean_or_dash(lst):
            return f"{(sum(lst)/len(lst)):.2f}" if len(lst) > 0 else "-"

        mean_len = mean_or_dash(len_vals)
        mean_area = mean_or_dash(area_vals)
        mean_pixels = mean_or_dash(pixel_vals)
        total_pixels = sum(pixel_vals) if pixel_vals else 0

        self.results_box.insert(tk.END, f"Média comprimento (cm): {mean_len}\n")
        self.results_box.insert(tk.END, f"Média área (cm²): {mean_area}\n")
        self.results_box.insert(tk.END, f"Média pixels (por medição de área): {mean_pixels}\n")
        self.results_box.insert(tk.END, f"Total pixels (soma de todas medições de área): {total_pixels}\n")
        self.results_box.insert(tk.END, f"Número total de medições: {len(self.results_data)}\n")

    # ---------------- salvar em XLSX (com formatação simples) ----------------
    def save_results(self):
        if not HAS_OPENPYXL:
            self.info_box.insert(tk.END, "openpyxl não instalado — não é possível salvar em XLSX.\n")
            return
        if not self.results_data:
            self.info_box.insert(tk.END, "Nenhum dado para salvar.\n")
            return

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                                 filetypes=[("Excel files", "*.xlsx")],
                                                 title="Salvar resultados como")
        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Resultados"

            headers = ["#", "Data/Hora", "Tipo", "Comprimento (cm)", "Área (cm²)", "Pixels",
                       "HSV_lower", "HSV_upper", "RGB_lower_from_HSV", "RGB_upper_from_HSV",
                       "RGB_min_escuro", "RGB_max_escuro", "RGB_min_claro", "RGB_max_claro",
                       "Tolerância (%)"]
            ws.append(headers)

            for i, r in enumerate(self.results_data, start=1):
                ws.append([
                    i,
                    r["timestamp"],
                    r["Tipo"],
                    r["Comprimento_cm"] if r["Comprimento_cm"] is not None else "",
                    r["Área_cm2"] if r["Área_cm2"] is not None else "",
                    r["Pixels"] if r["Pixels"] is not None else "",
                    str(r.get("HSV_lower", "")),
                    str(r.get("HSV_upper", "")),
                    str(r.get("RGB_lower_from_HSV", "")),
                    str(r.get("RGB_upper_from_HSV", "")),
                    str(r.get("RGB_min_escuro", "")),
                    str(r.get("RGB_max_escuro", "")),
                    str(r.get("RGB_min_claro", "")),
                    str(r.get("RGB_max_claro", "")),
                    r.get("Tolerance_percent", "")
                ])

            # formatação do cabeçalho
            header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # ajustar largura das colunas
            widths = [5, 20, 25, 18, 15, 12, 18, 18, 18, 18, 18, 18, 18, 18, 12]
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = w

            wb.save(file_path)
            self.info_box.insert(tk.END, f"Resultados salvos em: {file_path}\n")
        except Exception as e:
            self.info_box.insert(tk.END, f"Erro ao salvar arquivo: {e}\n")


if __name__ == "__main__":
    root = tk.Tk()
    app = GreenAreaDetector(root)
    root.mainloop()
