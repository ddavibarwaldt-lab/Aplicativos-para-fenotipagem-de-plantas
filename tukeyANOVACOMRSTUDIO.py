import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import os
import random
import subprocess
import glob
from openpyxl import Workbook

# Template do script R com apenas ANOVA e Tukey (sem MANOVA)
R_SCRIPT_TEMPLATE = '''
# Instalar pacotes se não estiverem instalados
if(!require(readxl)) install.packages("readxl")
if(!require(agricolae)) install.packages("agricolae")
if(!require(openxlsx)) install.packages("openxlsx")

library(readxl)
library(agricolae)
library(openxlsx)

# ----- ALTERE AQUI O CAMINHO DO SEU ARQUIVO -----
caminho_arquivo <- {caminho_arquivo}

# Lê os dados
dados <- read_excel(caminho_arquivo)

# Padroniza nomes das colunas
colnames(dados) <- toupper(iconv(colnames(dados), from="UTF-8", to="ASCII//TRANSLIT"))

# Visualiza as primeiras linhas
head(dados)

# Fator para ANOVA
fator <- dados$ESPECIE

# Lista para salvar resultados
resultados <- list()

# Variáveis numéricas a partir da 3ª coluna
col_vars <- colnames(dados)[3:ncol(dados)]

# Função para gerar tabela ANOVA formatada
gera_tabela_anova <- function(anova_obj, fator_nome) {{
  tab <- anova_obj[[1]]
  fv <- c(fator_nome, "Residuos", "Total")
  gl <- c(tab[1, "Df"], tab[2, "Df"], sum(tab[1:2, "Df"]))
  sq <- c(tab[1, "Sum Sq"], tab[2, "Sum Sq"], sum(tab[1:2, "Sum Sq"]))
  qm <- c(tab[1, "Mean Sq"], tab[2, "Mean Sq"], NA)
  fcalc <- c(tab[1, "F value"], NA, NA)
  ftab <- c(qf(0.95, gl[1], gl[2]), NA, NA)
  pval <- c(tab[1, "Pr(>F)"], NA, NA)
  df <- data.frame(FV=fv, GL=gl, SQ=round(sq,4), QM=round(qm,4),
                   FCALC=round(fcalc,4), FTAB=round(ftab,4), PVALOR=round(pval,4),
                   stringsAsFactors=FALSE)
  return(df)
}}

# ---- ANOVAS + TUKEY ----
for (var in col_vars) {{
  resposta <- as.numeric(dados[[var]])
  modelo <- aov(resposta ~ fator)
  anova_res <- summary(modelo)
  tabela_anova <- gera_tabela_anova(anova_res, "Tratamento")
  
  cat("\\n--- ANALISE da variavel:", var, "---\\n")
  print(tabela_anova)
  
  p_valor <- anova_res[[1]][["Pr(>F)"]][1]
  cat("p-valor ANOVA =", p_valor, "\\n")
  
  if (p_valor < 0.05) {{
    cat("Significativo! Fazendo teste de Tukey...\\n")
    tukey <- HSD.test(modelo, "fator", group=TRUE)
    medias_letras <- tukey$groups
    
    resumo <- aggregate(resposta, by=list(Grupo=fator), 
                        FUN=function(x) c(media=mean(x, na.rm=TRUE), 
                                          ep=sd(x, na.rm=TRUE)/sqrt(length(x))))
    resumo <- do.call(data.frame, resumo)
    colnames(resumo) <- c("Tratamento", "Media", "EP")
    
    tabela_final <- merge(resumo, 
                          data.frame(Tratamento=rownames(medias_letras), 
                                     Letra=medias_letras$groups), 
                          by="Tratamento")
    
    tabela_final$`Media ± EP` <- paste0(round(tabela_final$Media, 2), 
                                        " ± ", 
                                        round(tabela_final$EP, 2))
    tabela_final <- tabela_final[, c("Tratamento", "Media ± EP", "Letra")]
    
    print(tabela_final)
    resultados[[var]] <- list(anova = tabela_anova, tukey = tabela_final)
  }} else {{
    cat("Nao significativo, pulando Tukey.\\n")
    resultados[[var]] <- list(anova = tabela_anova)
  }}
}}

# ---- SALVAR RESULTADOS EM EXCEL ----
wb <- createWorkbook()

# Salvar ANOVA + Tukey
for (var in names(resultados)) {{
  addWorksheet(wb, var)
  writeData(wb, var, resultados[[var]]$anova, startRow=1, colNames=TRUE)
  if (!is.null(resultados[[var]]$tukey)) {{
    writeData(wb, var, resultados[[var]]$tukey, startRow = nrow(resultados[[var]]$anova) + 4, colNames = TRUE)
  }}
}}

caminho_resultados <- {caminho_resultados}
saveWorkbook(wb, caminho_resultados, overwrite = TRUE)
cat("\\nAnalise finalizada! Resultados salvos em:\\n", caminho_resultados, "\\n")
'''


def gerar_script(caminho):
    pasta = os.path.dirname(caminho).replace("\\", "/")
    nome_base = os.path.splitext(os.path.basename(caminho))[0]
    caminho_r = caminho.replace("\\", "/")
    caminho_resultados = os.path.join(pasta, f"{nome_base}_resultados.xlsx").replace("\\", "/")
    return R_SCRIPT_TEMPLATE.format(
        caminho_arquivo=f'"{caminho_r}"',
        caminho_resultados=f'"{caminho_resultados}"'
    ), caminho_resultados


def selecionar_arquivo():
    global caminho_dados, script_r, caminho_resultados
    caminho_dados = filedialog.askopenfilename(
        title="Selecione o arquivo Excel com os dados",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    if caminho_dados:
        script_r, caminho_resultados = gerar_script(caminho_dados)
        text_area.delete('1.0', tk.END)
        text_area.insert(tk.END, script_r)


def gerar_arquivo_exemplo():
    especies = ["Batata", "Melao", "Abobora", "Tomate"]
    parametros = ["Altura", "Peso", "Comprimento", "Grau_brix"]

    wb = Workbook()
    ws = wb.active
    ws.title = "Dados"
    ws.append(["Repeticao", "Especie"] + parametros)

    for especie in especies:
        for rep in range(1, 5):
            linha = [rep, especie]
            for param in parametros:
                if param == "Altura":
                    valor = round(random.uniform(10, 100), 1)
                elif param == "Peso":
                    valor = round(random.uniform(0.1, 5), 2)
                elif param == "Comprimento":
                    valor = round(random.uniform(5, 30), 1)
                elif param == "Grau_brix":
                    valor = round(random.uniform(5, 15), 1)
                linha.append(valor)
            ws.append(linha)

    caminho = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel", "*.xlsx")],
        title="Salvar arquivo de exemplo"
    )
    if caminho:
        wb.save(caminho)
        messagebox.showinfo("Sucesso", f"Arquivo de exemplo salvo em:\n{caminho}")


# ====== DETECTAR Rscript.exe ======
def encontrar_rscript():
    try:
        subprocess.run(["Rscript", "--version"], capture_output=True, text=True, check=True)
        return "Rscript"
    except Exception:
        pass

    caminhos = glob.glob(r"C:\\Program Files\\R\\R-*\\bin\\Rscript.exe")
    if caminhos:
        return caminhos[-1]
    caminhos = glob.glob(r"C:\\Program Files (x86)\\R\\R-*\\bin\\Rscript.exe")
    if caminhos:
        return caminhos[-1]
    return None


# ====== RODAR SCRIPT R ======
def rodar_script_r():
    global script_r, caminho_resultados
    if not script_r:
        messagebox.showwarning("Aviso", "Selecione primeiro um arquivo Excel para gerar o script!")
        return

    rscript_path = encontrar_rscript()
    if not rscript_path:
        messagebox.showerror("Erro", "Não foi possível localizar o Rscript.exe.\nVerifique se o R está instalado.")
        return

    caminho_r = os.path.join(os.getcwd(), "script_temp.R")
    with open(caminho_r, "w", encoding="utf-8") as f:
        f.write(script_r)

    try:
        resultado = subprocess.run([rscript_path, caminho_r], capture_output=True, text=True, check=True)
        text_area.delete('1.0', tk.END)
        text_area.insert(tk.END, resultado.stdout)
        messagebox.showinfo("Sucesso", f"Script R executado com sucesso!\nResultados salvos em:\n{caminho_resultados}")
    except subprocess.CalledProcessError as e:
        text_area.delete('1.0', tk.END)
        text_area.insert(tk.END, e.stdout + "\\n" + e.stderr)
        messagebox.showerror("Erro", "Falha ao executar o script R.")


# ====== INTERFACE TKINTER ======
root = tk.Tk()
root.title("Gerador de Script R para ANOVA e Tukey")
root.geometry("900x700")

script_r = None
caminho_dados = None
caminho_resultados = None

btn_exemplo = tk.Button(root, text="Gerar arquivo de exemplo", command=gerar_arquivo_exemplo)
btn_exemplo.pack(pady=5)

btn_selecionar = tk.Button(root, text="Selecionar arquivo Excel", command=selecionar_arquivo)
btn_selecionar.pack(pady=5)

btn_rodar = tk.Button(root, text="Rodar Script R", command=rodar_script_r)
btn_rodar.pack(pady=5)

text_area = scrolledtext.ScrolledText(root, width=100, height=35)
text_area.pack(padx=10, pady=10)

root.mainloop()
