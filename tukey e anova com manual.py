import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
import os
import pandas as pd
from openpyxl import Workbook

# ---------- Template R atualizado com asteriscos ----------
R_SCRIPT_TEMPLATE = '''
if(!require(readxl)) install.packages("readxl")
if(!require(agricolae)) install.packages("agricolae")
if(!require(openxlsx)) install.packages("openxlsx")

library(readxl)
library(agricolae)
library(openxlsx)

caminho_arquivo <- {caminho_arquivo}

dados <- read_excel(caminho_arquivo)
colnames(dados) <- toupper(iconv(colnames(dados), from="UTF-8", to="ASCII//TRANSLIT"))
head(dados)

fator <- dados$ESPECIE
{repeticao_definicao}

resultados <- list()
col_vars <- colnames(dados)[3:ncol(dados)]

gera_tabela_anova <- function(anova_obj, fator_nome) {{
  tab <- anova_obj[[1]]
  fv <- c(fator_nome, "Resíduos", "Total")
  gl <- c(tab[1, "Df"], tab[2, "Df"], sum(tab[1:2, "Df"]))
  sq <- c(tab[1, "Sum Sq"], tab[2, "Sum Sq"], sum(tab[1:2, "Sum Sq"]))
  qm <- c(tab[1, "Mean Sq"], tab[2, "Mean Sq"], NA)
  fcalc <- c(tab[1, "F value"], NA, NA)
  ftab <- c(qf(0.95, gl[1], gl[2]), NA, NA)
  pval <- c(tab[1, "Pr(>F)"], NA, NA)
  df <- data.frame(FV=fv, GL=gl, SQ=round(sq,4), QM=round(qm,4),
                   FCALC=round(fcalc,4), FTAB=round(ftab,4), PVALOR=round(pval,4),
                   stringsAsFactors=FALSE)
  return(df)
}}

for (var in col_vars) {{
  resposta <- as.numeric(dados[[var]])
  {modelo_anova}

  anova_res <- summary(modelo)
  tabela_anova <- gera_tabela_anova(anova_res, "Tratamento")
  
  cat("\\n--- ANÁLISE da variável:", var, "---\\n")
  print(tabela_anova)
  
  p_valor <- anova_res[[1]][["Pr(>F)"]][1]
  cat("p-valor ANOVA =", p_valor, "\\n")
  
  if (p_valor < 0.05) {{
    cat("Significativo! Fazendo teste de Tukey...\\n")
    tukey <- HSD.test(modelo, "fator", group=TRUE)
    medias_letras <- tukey$groups
    
    resumo <- aggregate(resposta, by=list(Grupo=fator), 
                        FUN=function(x) c(media=mean(x, na.rm=TRUE), 
                                          ep=sd(x, na.rm=TRUE)/sqrt(length(x))))
    resumo <- do.call(data.frame, resumo)
    colnames(resumo) <- c("Tratamento", "Media", "EP")
    
    tabela_final <- merge(resumo, 
                          data.frame(Tratamento=rownames(medias_letras), 
                                     Letra=medias_letras$groups), 
                          by="Tratamento")
    
    # --- ADICIONA ASTERISCOS NA COLUNA Media ± EP ---
    sinal <- ""
    if(p_valor < 0.05) sinal <- "*"
    if(p_valor < 0.01) sinal <- "**"
    if(p_valor < 0.001) sinal <- "***"
    tabela_final$`Media ± EP` <- paste0(round(tabela_final$Media,2),
                                        " ± ", round(tabela_final$EP,2),
                                        " ", sinal)
    
    tabela_final <- tabela_final[, c("Tratamento", "Media ± EP", "Letra")]
    print(tabela_final)
    resultados[[var]] <- list(anova = tabela_anova, tukey = tabela_final)
  }} else {{
    cat("Não significativo, pulando Tukey.\\n")
    resultados[[var]] <- list(anova = tabela_anova)
  }}
}}

# ---- SALVAR RESULTADOS EM EXCEL ----
wb <- createWorkbook()
for (var in names(resultados)) {{
  addWorksheet(wb, var)
  writeData(wb, var, resultados[[var]]$anova, startRow=1, colNames=TRUE)
  if (!is.null(resultados[[var]]$tukey)) {{
    writeData(wb, var, resultados[[var]]$tukey, 
              startRow = nrow(resultados[[var]]$anova) + 4, colNames = TRUE)
  }}
}}

caminho_resultados <- {caminho_resultados}
saveWorkbook(wb, caminho_resultados, overwrite = TRUE)
cat("\\nAnálise finalizada! Resultados salvos em:\\n", caminho_resultados, "\\n")
'''

# ---------- Funções auxiliares ----------
def gerar_script(caminho, delineamento):
    pasta = os.path.dirname(caminho).replace("\\", "/")
    nome_base = os.path.splitext(os.path.basename(caminho))[0]
    caminho_r = caminho.replace("\\", "/")
    caminho_resultados = os.path.join(pasta, f"{nome_base}_resultados.xlsx").replace("\\", "/")

    if delineamento == "DIC":
        repeticao_definicao = ""
        modelo_anova = "modelo <- aov(resposta ~ fator)"
    else:
        repeticao_definicao = "repeticao <- dados$REPETICAO  # Certifique-se que esta coluna existe"
        modelo_anova = "modelo <- aov(resposta ~ fator + repeticao)"

    return R_SCRIPT_TEMPLATE.format(
        caminho_arquivo=f'"{caminho_r}"',
        caminho_resultados=f'"{caminho_resultados}"',
        repeticao_definicao=repeticao_definicao,
        modelo_anova=modelo_anova
    )

def validar_excel(caminho, delineamento):
    try:
        df = pd.read_excel(caminho)
    except Exception as e:
        messagebox.showerror("Erro", f"Não foi possível ler o arquivo:\n{e}")
        return False

    colunas = list(df.columns)
    erros = []

    if len(colunas) < 2:
        erros.append("O arquivo deve ter pelo menos 2 colunas: ID e ESPECIE")
    else:
        if colunas[1].upper() != "ESPECIE":
            erros.append("A segunda coluna deve se chamar 'ESPECIE'")

    if delineamento.get() == "DBC" and "REPETICAO" not in [c.upper() for c in colunas]:
        erros.append("Para DBC, a coluna 'REPETICAO' é obrigatória")

    idx_inicio = 2 if delineamento.get() == "DIC" else 3
    for c in colunas[idx_inicio:]:
        if not pd.api.types.is_numeric_dtype(df[c]):
            erros.append(f"A coluna '{c}' deve conter valores numéricos")

    if erros:
        messagebox.showerror("Arquivo inválido", "O arquivo não atende aos requisitos:\n- " + "\n- ".join(erros))
        return False

    return True

def selecionar_arquivo(delineamento):
    caminho = filedialog.askopenfilename(
        title="Selecione o arquivo Excel",
        filetypes=[("Arquivos Excel", "*.xlsx *.xls")]
    )
    if not caminho:
        return
    if not validar_excel(caminho, delineamento):
        return
    script_r = gerar_script(caminho, delineamento.get())
    text_area.delete('1.0', tk.END)
    text_area.insert(tk.END, script_r)

def gerar_exemplo_excel():
    caminho_exemplo = filedialog.asksaveasfilename(
        title="Salvar arquivo de exemplo Excel",
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        initialfile="exemplo_DIC_DBC.xlsx"
    )
    if not caminho_exemplo:
        return

    wb = Workbook()

    # ----- Planilha DIC -----
    ws_dic = wb.active
    ws_dic.title = "DIC"
    ws_dic.append(["ID", "ESPECIE", "Altura", "Peso", "Raiz", "Produção"])
    # Altura e Produção significativas; Peso e Raiz não
    dados_dic = [
        [1, "CultivarA", 12.5, 3.4, 5.1, 40.2],
        [2, "CultivarA", 11.8, 3.2, 5.0, 39.8],
        [3, "CultivarA", 13.0, 3.5, 5.2, 41.0],
        [4, "CultivarB", 17.2, 3.6, 5.3, 48.5],
        [5, "CultivarB", 18.1, 3.7, 5.4, 49.1],
        [6, "CultivarB", 17.8, 3.4, 5.5, 47.9],
    ]
    for linha in dados_dic:
        ws_dic.append(linha)

    # ----- Planilha DBC -----
    ws_dbc = wb.create_sheet("DBC")
    ws_dbc.append(["ID", "ESPECIE", "REPETICAO", "Altura", "Peso", "Raiz", "Produção"])
    dados_dbc = [
        [1, "CultivarA", 1, 12.5, 3.4, 5.1, 40.2],
        [2, "CultivarA", 2, 11.8, 3.2, 5.0, 39.8],
        [3, "CultivarA", 3, 13.0, 3.5, 5.2, 41.0],
        [4, "CultivarB", 1, 17.2, 3.6, 5.3, 48.5],
        [5, "CultivarB", 2, 18.1, 3.7, 5.4, 49.1],
        [6, "CultivarB", 3, 17.8, 3.5, 5.5, 47.9],
    ]
    for linha in dados_dbc:
        ws_dbc.append(linha)

    wb.save(caminho_exemplo)
    messagebox.showinfo(
        "Exemplo gerado",
        f"Arquivo de exemplo salvo em:\n{caminho_exemplo}\n\n"
        "Variáveis com ANOVA significativa (devem gerar Tukey): Altura, Produção.\n"
        "Variáveis não significativas (sem Tukey): Peso, Raiz.\n\n"
        "Níveis de significância: * p<0.05, ** p<0.01, *** p<0.001"
    )

# ---------- INTERFACE Tkinter ----------
root = tk.Tk()
root.title("Gerador de Script R para ANOVA e Tukey")
root.geometry("1000x950")

cabecalho = (
    "INSTRUÇÕES DE FORMATO DO ARQUIVO EXCEL:\n"
    "1ª coluna: Identificador da observação (mesmo para DIC e DBC)\n"
    "2ª coluna: ESPECIE (nome do cultivar, fator principal)\n"
    "Demais colunas: Variáveis/ parâmetros a serem analisados\n"
    "Para DBC, inclua também uma coluna chamada REPETICAO indicando o bloco de cada observação\n\n"
    "💡 No arquivo de exemplo:\n"
    " - 'Altura' e 'Produção' geram Tukey (significativas)\n"
    " - 'Peso' e 'Raiz' não geram Tukey\n"
    " - Níveis de significância indicados por: * ** ***\n"
)
tk.Label(root, text=cabecalho, justify=tk.LEFT, fg="blue").pack(pady=10, padx=10, anchor="w")

frame_exemplo = tk.Frame(root, relief=tk.RIDGE, bd=2)
frame_exemplo.pack(padx=10, pady=5, fill="x")
tk.Label(frame_exemplo, text="Exemplo de formatação de tabela:", font=("Arial", 10, "bold")).pack(anchor="w", padx=5)
exemplo_texto = (
    "DIC:\n"
    "ID | ESPECIE  | Altura | Peso | Raiz | Produção\n"
    "1  | CultivarA| 12.5   | 3.4  | 5.1  | 40.2\n"
    "2  | CultivarA| 11.8   | 3.2  | 5.0  | 39.8\n"
    "3  | CultivarA| 13.0   | 3.5  | 5.2  | 41.0\n"
    "4  | CultivarB| 17.2   | 3.6  | 5.3  | 48.5\n"
    "5  | CultivarB| 18.1   | 3.7  | 5.4  | 49.1\n"
    "6  | CultivarB| 17.8   | 3.4  | 5.5  | 47.9\n"
)
tk.Label(frame_exemplo, text=exemplo_texto, justify=tk.LEFT, font=("Courier", 10), fg="darkgreen").pack(padx=5, pady=5, anchor="w")

delineamento = tk.StringVar(value="DIC")
frame_top = tk.Frame(root)
frame_top.pack(pady=5, fill="x", padx=10)
tk.Label(frame_top, text="Escolha o delineamento experimental:").pack(side=tk.LEFT, padx=5)
tk.Radiobutton(frame_top, text="DIC", variable=delineamento, value="DIC").pack(side=tk.LEFT)
tk.Radiobutton(frame_top, text="DBC", variable=delineamento, value="DBC").pack(side=tk.LEFT)
btn_selecionar = tk.Button(frame_top, text="Selecionar arquivo Excel", 
                            command=lambda: selecionar_arquivo(delineamento))
btn_selecionar.pack(side=tk.LEFT, padx=10)
btn_exemplo = tk.Button(frame_top, text="Gerar exemplo Excel", command=gerar_exemplo_excel)
btn_exemplo.pack(side=tk.LEFT, padx=10)

text_area = scrolledtext.ScrolledText(root, width=120, height=40)
text_area.pack(padx=10, pady=10)

root.mainloop()
